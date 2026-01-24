#!/usr/bin/env python3
"""
엑셀 리그 기록 파일을 JSON으로 변환하는 스크립트
사용법: python3 convert_excel_to_json.py <엑셀파일경로> <시즌코드>
예시: python3 convert_excel_to_json.py "/Users/user/Downloads/2026-01 리그 기록.xlsx" 202601
"""

import json
import sys
import re
import openpyxl
from pathlib import Path


def parse_team_from_rows(ws, start_row, end_row):
    """전체득점 시트에서 선수 데이터를 파싱"""
    players = []
    current_team = None

    for row_idx in range(start_row, end_row + 1):
        row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]

        # 팀 정보 (A열)
        if row[0] is not None:
            current_team = row[0]

        # 선수명 (B열)과 번호 (C열)
        player_name = row[1]
        player_number = row[2]

        if player_name is None or player_number is None:
            continue

        # 참석수 (S열, 인덱스 18)
        attendance = row[18] if len(row) > 18 and row[18] is not None else 0
        attendance = int(attendance) if attendance else 0

        # 총득점 (T열, 인덱스 19)
        total_score = row[19] if len(row) > 19 and row[19] is not None else 0
        total_score = int(total_score) if total_score else 0

        # 평균득점 (V열, 인덱스 21)
        avg_score = row[21] if len(row) > 21 and row[21] is not None else 0
        avg_score = float(avg_score) if avg_score else 0.0

        players.append({
            'team': current_team,
            'name': player_name,
            'number': int(player_number),
            'attendance': attendance,
            'total_score': total_score,
            'avg_score': round(avg_score, 1)
        })

    return players


def parse_additional_stats(ws):
    """부가기록 계산 시트에서 부가기록 데이터를 파싱"""
    stats = {}

    for row_idx in range(3, ws.max_row + 1):
        row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]

        player_name = row[0]
        player_number = row[1]

        if player_name is None or player_number is None:
            continue

        # 누적 (C~G열, 인덱스 2~6): 리바운드, 어시스트, 스틸, 블록, 3점슛
        rebound_total = int(row[2]) if row[2] is not None else 0
        assist_total = int(row[3]) if row[3] is not None else 0
        steal_total = int(row[4]) if row[4] is not None else 0
        block_total = int(row[5]) if row[5] is not None else 0
        three_pt_total = int(row[6]) if row[6] is not None else 0

        # 평균 (H~L열, 인덱스 7~11): 리바운드, 어시스트, 스틸, 블록, 3점슛
        rebound_avg = float(row[7]) if row[7] is not None else 0.0
        assist_avg = float(row[8]) if row[8] is not None else 0.0
        steal_avg = float(row[9]) if row[9] is not None else 0.0
        block_avg = float(row[10]) if row[10] is not None else 0.0
        three_pt_avg = float(row[11]) if row[11] is not None else 0.0

        key = (player_name, int(player_number))
        stats[key] = {
            '리바운드': {'누적': rebound_total, '평균': round(rebound_avg, 1)},
            '어시스트': {'누적': assist_total, '평균': round(assist_avg, 1)},
            '스틸': {'누적': steal_total, '평균': round(steal_avg, 1)},
            '블록': {'누적': block_total, '평균': round(block_avg, 1)},
            '3점슛': {'누적': three_pt_total, '평균': round(three_pt_avg, 1)}
        }

    return stats


def count_rounds(ws):
    """전체득점 시트 헤더에서 라운드 수를 계산

    정확히 'N라운드' 형식의 컬럼만 카운트 (예: 1라운드, 2라운드, ...)
    '라운드 합계' 같은 컬럼은 제외
    """
    header = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    round_count = 0

    for cell in header:
        if cell and re.match(r'^\d+라운드$', str(cell).strip()):
            round_count += 1

    return round_count


def get_current_round(ws):
    """전체득점 시트에서 현재 진행된 라운드를 계산

    2행(첫 번째 선수 데이터)의 각 라운드 컬럼에 값이 있는지 확인하여
    값이 있는 마지막 라운드를 현재 라운드로 판단
    """
    header = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    data_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]

    # 라운드 컬럼 인덱스와 라운드 번호 매핑
    round_columns = []
    for i, cell in enumerate(header):
        if cell and '라운드' in str(cell):
            match = re.search(r'(\d+)라운드', str(cell))
            if match:
                round_num = int(match.group(1))
                round_columns.append((i, round_num))

    # 값이 있는 마지막 라운드 찾기
    current_round = 0
    for col_idx, round_num in round_columns:
        if col_idx < len(data_row) and data_row[col_idx] is not None:
            current_round = round_num

    return current_round


def parse_record(record_str):
    """'1승 1패' 형식의 문자열을 파싱하여 승/패 수를 반환"""
    wins = 0
    losses = 0
    if record_str:
        win_match = re.search(r'(\d+)승', str(record_str))
        loss_match = re.search(r'(\d+)패', str(record_str))
        if win_match:
            wins = int(win_match.group(1))
        if loss_match:
            losses = int(loss_match.group(1))
    return wins, losses


def extract_name(text, prefix):
    """'MOM: 권인회' 형식에서 이름을 추출"""
    if text and prefix in str(text):
        clean = re.sub(r'[👑✌️🏀]', '', str(text)).strip()
        match = re.search(prefix + r'\s*(.+)', clean)
        if match:
            return match.group(1).strip()
    return None


def extract_scorer(text):
    """'오늘 득점왕: 강재훈(66점)' 형식에서 이름과 점수를 추출"""
    if text and '득점왕' in str(text):
        clean = re.sub(r'[👑✌️🏀]', '', str(text)).strip()
        match = re.search(r'득점왕:\s*(\S+)\((\d+)점\)', clean)
        if match:
            return match.group(1), int(match.group(2))
    return None, None


def parse_gbl_standings(wb):
    """GBL 승점 시트에서 팀 순위와 어워드 데이터를 파싱"""
    if 'GBL 승점' not in wb.sheetnames:
        return None

    ws = wb['GBL 승점']
    rounds = []

    for row in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=row, column=1).value
        if cell_a and '라운드 리그 누적 결과' in str(cell_a):
            match = re.search(r'(\d+)라운드', str(cell_a))
            if match:
                current_round = int(match.group(1))

                round_info = {
                    'round': current_round,
                    'teams': [],
                    'awards': {}
                }

                # 팀 데이터 읽기 (행 +2 ~ +4)
                for team_row in range(row + 2, row + 5):
                    team_name = ws.cell(row=team_row, column=1).value
                    win_lose = ws.cell(row=team_row, column=2).value
                    points = ws.cell(row=team_row, column=3).value

                    if team_name and team_name in ['A팀', 'B팀', 'C팀']:
                        wins, losses = parse_record(win_lose)
                        team_code = team_name[0]
                        round_info['teams'].append({
                            'team': team_code,
                            'name': team_name,
                            'record': str(win_lose) if win_lose else '0승 0패',
                            'wins': wins,
                            'losses': losses,
                            'points': float(points) if points else 0
                        })

                # 어워드 찾기
                for search_row in range(row, min(row + 15, ws.max_row + 1)):
                    for col in range(1, ws.max_column + 1):
                        cell_val = ws.cell(row=search_row, column=col).value
                        if cell_val:
                            cell_str = str(cell_val)

                            # MOM
                            if 'MOM:' in cell_str:
                                mom = extract_name(cell_str, 'MOM:')
                                if mom:
                                    round_info['awards']['mom'] = mom

                            # 더블더블
                            if '더블더블:' in cell_str:
                                dd = extract_name(cell_str, '더블더블:')
                                if dd:
                                    round_info['awards']['doubleDouble'] = dd

                            # 득점왕
                            if '득점왕:' in cell_str:
                                scorer, pts = extract_scorer(cell_str)
                                if scorer:
                                    round_info['awards']['topScorer'] = {
                                        'name': scorer,
                                        'points': pts
                                    }

                rounds.append(round_info)

    return rounds


def generate_metadata(season_name, total_rounds, rounds_data, current_round):
    """메타데이터 JSON 생성

    Args:
        season_name: 시즌 이름 (예: "2026년 1월")
        total_rounds: 총 라운드 수
        rounds_data: GBL 승점 시트에서 파싱한 라운드별 데이터
        current_round: 전체득점 시트 기준 현재 진행된 라운드
    """
    if not rounds_data:
        return None

    # 현재 라운드에 해당하는 데이터 찾기
    latest_round = None
    for rd in rounds_data:
        if rd['round'] == current_round:
            latest_round = rd
            break

    # 못 찾으면 가장 최신 라운드 사용
    if latest_round is None:
        latest_round = rounds_data[-1]
    standings = []

    for team_data in latest_round.get('teams', []):
        standings.append({
            'team': team_data['team'],
            'name': team_data['name'],
            'wins': team_data['wins'],
            'losses': team_data['losses'],
            'points': team_data['points']
        })

    # 승점 순으로 정렬
    standings = sorted(standings, key=lambda x: x['points'], reverse=True)

    metadata = {
        'season': season_name,
        'currentRound': current_round,
        'totalRounds': total_rounds,
        'standings': standings,
        'roundHistory': rounds_data
    }

    return metadata


def convert_excel_to_json(excel_path, season_code):
    """엑셀 파일을 JSON으로 변환"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # 전체득점 시트에서 선수 기본 정보 추출
    ws_score = wb['전체득점']

    # 라운드 수 계산
    total_rounds = count_rounds(ws_score)

    # 선수 데이터 추출 (2행부터 데이터 시작, 마지막 행까지)
    players_basic = parse_team_from_rows(ws_score, 2, ws_score.max_row)

    # 부가기록 시트에서 추가 통계 추출
    ws_stats = wb['부가기록 계산']
    additional_stats = parse_additional_stats(ws_stats)

    # 데이터 병합
    players_list = []
    for player in players_basic:
        key = (player['name'], player['number'])
        stats = additional_stats.get(key, {
            '리바운드': {'누적': 0, '평균': 0.0},
            '어시스트': {'누적': 0, '평균': 0.0},
            '스틸': {'누적': 0, '평균': 0.0},
            '블록': {'누적': 0, '평균': 0.0},
            '3점슛': {'누적': 0, '평균': 0.0}
        })

        players_list.append({
            '번호': player['number'],
            '팀': player['team'],
            '선수명': player['name'],
            '득점': {
                '누적득점': player['total_score'],
                '평균득점': player['avg_score']
            },
            '출석': player['attendance'],
            '부가기록': {
                '어시스트': stats['어시스트'],
                '리바운드': stats['리바운드'],
                '스틸': stats['스틸'],
                '블록': stats['블록'],
                '3점슛': stats['3점슛']
            }
        })

    # 시즌 이름 생성 (202601 -> "2026년 1월")
    year = season_code[:4]
    month = int(season_code[4:6])
    season_name = f"{year}년 {month}월"

    result = {
        '시즌': season_name,
        '총라운드': total_rounds,
        '총선수수': len(players_list),
        '선수목록': players_list
    }

    return result


def main():
    if len(sys.argv) < 3:
        print("사용법: python3 convert_excel_to_json.py <엑셀파일경로> <시즌코드>")
        print("예시: python3 convert_excel_to_json.py '/Users/user/Downloads/2026-01 리그 기록.xlsx' 202601")
        sys.exit(1)

    excel_path = sys.argv[1]
    season_code = sys.argv[2]

    if not Path(excel_path).exists():
        print(f"오류: 파일을 찾을 수 없습니다 - {excel_path}")
        sys.exit(1)

    # 엑셀 파일 로드
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # GBL 승점 시트에서 현재 라운드 가져오기 (가장 신뢰할 수 있는 소스)
    rounds_data = parse_gbl_standings(wb)
    if rounds_data:
        # 가장 최신 라운드 사용
        current_round = max(rd['round'] for rd in rounds_data)
    else:
        # GBL 승점 시트가 없으면 전체득점 시트 기준 폴백
        ws_score = wb['전체득점']
        current_round = get_current_round(ws_score)

    # 선수 통계 JSON 생성
    result = convert_excel_to_json(excel_path, season_code)

    # 선수 통계 JSON 저장
    output_path = Path(__file__).parent / f"league_stats_{season_code}.json"
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"선수 통계 저장: {output_path}")
    print(f"  시즌: {result['시즌']}")
    print(f"  총 라운드: {result['총라운드']}")
    print(f"  현재 라운드: {current_round}")
    print(f"  총 선수 수: {result['총선수수']}")

    # GBL 승점 시트 데이터로 메타데이터 생성
    if rounds_data:
        metadata = generate_metadata(result['시즌'], result['총라운드'], rounds_data, current_round)

        if metadata:
            # 메타데이터 JSON 저장
            metadata_path = Path(__file__).parent / f"league_metadata_{season_code}.json"
            with open(metadata_path, 'w', encoding='utf-8') as f:
                json.dump(metadata, f, ensure_ascii=False, indent=2)

            print(f"\n메타데이터 저장: {metadata_path}")

            # manifest 업데이트
            manifest_path = Path(__file__).parent / "metadata_manifest.json"
            manifest = {"seasons": []}
            if manifest_path.exists():
                with open(manifest_path, 'r', encoding='utf-8') as f:
                    manifest = json.load(f)

            if season_code not in manifest["seasons"]:
                manifest["seasons"].append(season_code)

            with open(manifest_path, 'w', encoding='utf-8') as f:
                json.dump(manifest, f, ensure_ascii=False, indent=2)

            print(f"매니페스트 업데이트: {manifest_path}")
            print(f"  현재 라운드: {metadata['currentRound']}/{metadata['totalRounds']}")
            print(f"  팀 순위:")
            for i, team in enumerate(metadata['standings'], 1):
                print(f"    {i}위: {team['name']} ({team['wins']}승 {team['losses']}패, {team['points']}점)")

            # 현재 라운드에 해당하는 어워드 찾기
            current_round_data = None
            for rd in metadata['roundHistory']:
                if rd['round'] == current_round:
                    current_round_data = rd
                    break

            if current_round_data and current_round_data.get('awards'):
                print(f"  {current_round}라운드 어워드:")
                if current_round_data['awards'].get('mom'):
                    print(f"    MOM: {current_round_data['awards']['mom']}")
                if current_round_data['awards'].get('doubleDouble'):
                    print(f"    더블더블: {current_round_data['awards']['doubleDouble']}")
                if current_round_data['awards'].get('topScorer'):
                    scorer = current_round_data['awards']['topScorer']
                    print(f"    득점왕: {scorer['name']}({scorer['points']}점)")
    else:
        print("\n메타데이터: GBL 승점 시트를 찾을 수 없거나 데이터가 없습니다.")


if __name__ == '__main__':
    main()
